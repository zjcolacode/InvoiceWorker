"""报销单管理 API：上传发票明细清单核销 + 历史记录查询。"""
import base64
import csv
import email as email_lib
import imaplib
import io
import json
import logging
import os
import re
import tempfile
from datetime import date, datetime, timedelta, timezone
from email.header import decode_header
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Optional

import httpx
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from openpyxl import load_workbook
from sqlalchemy import or_
from sqlalchemy.orm import Session

try:
    import fitz  # PyMuPDF
except ImportError:  # pragma: no cover
    fitz = None

from app.core.config import settings
from app.core.database import get_db
from app.core.security import get_current_user
from app.models.email_message import EmailMessage
from app.models.invoice import Invoice
from app.models.invoice_detail import InvoiceDetail, InvoiceDetailUploadLog
from app.models.manual_match import ManualMatchRecord
from app.models.reimb_email import ReimbEmailConfig, ReimbEmailFetchLog, ReimbEmailMessage
from app.models.reimbursement import ReimbursementRecord
from app.models.reimbursement_application import ReimbursementApplication
from app.models.user import User
from app.schemas.reimbursement import (
    InvoiceDetailListResponse,
    InvoiceDetailResponse,
    ManualMatchRecognizeResult,
    ManualMatchRecordResponse,
    ReimbEmailConfigCreate,
    ReimbEmailConfigResponse,
    ReimbEmailConfigUpdate,
    ReimbEmailFetchLogResponse,
    ReimbEmailFetchRequest,
    ReimbEmailMessageResponse,
    ReimbEmailTestResult,
    ReimbursementApplicationCreate,
    ReimbursementRecordResponse,
    ReimbursementResult,
    UploadDetailResult,
    UploadLogListResponse,
    UploadLogResponse,
    VerifyResult,
)
from app.services.email_fetcher import decrypt_password, encrypt_password

logger = logging.getLogger(__name__)
router = APIRouter()


def _find_invoice_no_columns(header: list) -> tuple[int | None, int | None]:
    """在表头中查找「发票号码」和「数电发票号码」列的索引。

    返回 (发票号码列索引, 数电发票号码列索引)，未找到的为 None。
    """
    invoice_no_idx = None
    digital_no_idx = None
    for idx, cell in enumerate(header):
        if cell is None:
            continue
        name = str(cell).strip()
        if name == "数电发票号码":
            digital_no_idx = idx
        elif "发票号码" in name and invoice_no_idx is None:
            invoice_no_idx = idx
    return invoice_no_idx, digital_no_idx


def _extract_invoice_no(row: tuple, no_idx: int | None, digital_idx: int | None) -> str | None:
    """从一行中提取发票号码：优先取「发票号码」，为空则回退到「数电发票号码」。"""
    for idx in (no_idx, digital_idx):
        if idx is not None and idx < len(row) and row[idx] is not None:
            val = str(row[idx]).strip()
            if val:
                return val
    return None


def _parse_invoice_numbers_from_xlsx(content: bytes) -> list[str]:
    """从 xlsx 文件中提取发票号码（支持「发票号码」和「数电发票号码」）。"""
    wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel 文件没有活动工作表")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 文件为空")

    header = rows[0]
    no_idx, digital_idx = _find_invoice_no_columns(header)
    if no_idx is None and digital_idx is None:
        raise ValueError("未找到「发票号码」或「数电发票号码」列，请检查文件表头")

    invoice_nos = []
    for row in rows[1:]:
        val = _extract_invoice_no(row, no_idx, digital_idx)
        if val:
            invoice_nos.append(val)
    wb.close()
    return invoice_nos


def _parse_invoice_numbers_from_csv(content: bytes) -> list[str]:
    """从 csv 文件中提取「发票号码」列的所有值。"""
    # 尝试多种编码
    text = None
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb2312"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("无法识别文件编码，请使用 UTF-8 或 GBK 编码")

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise ValueError("CSV 文件为空")

    header = rows[0]
    no_idx, digital_idx = _find_invoice_no_columns(header)
    if no_idx is None and digital_idx is None:
        raise ValueError("未找到「发票号码」或「数电发票号码」列，请检查文件表头")

    invoice_nos = []
    for row in rows[1:]:
        # csv rows 是 list[str]，转为 tuple 以复用 _extract_invoice_no
        val = _extract_invoice_no(tuple(row), no_idx, digital_idx)
        if val:
            invoice_nos.append(val)
    return invoice_nos


# ============================================================
# 上传核销
# ============================================================
@router.post("/upload", response_model=ReimbursementResult)
async def upload_and_verify(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """上传发票明细清单（xlsx/csv），根据发票号码核销已有发票记录。"""
    filename = file.filename or "unknown"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext not in ("xlsx", "csv"):
        raise HTTPException(status_code=400, detail="仅支持 xlsx 和 csv 格式文件")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="文件内容为空")

    # 解析发票号码
    try:
        if ext == "xlsx":
            invoice_nos = _parse_invoice_numbers_from_xlsx(content)
        else:
            invoice_nos = _parse_invoice_numbers_from_csv(content)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # 去重（保留顺序）
    seen = set()
    unique_nos = []
    for no in invoice_nos:
        if no not in seen:
            seen.add(no)
            unique_nos.append(no)

    if not unique_nos:
        raise HTTPException(status_code=400, detail="文件中未解析到有效的发票号码")

    # 查询系统中匹配的发票
    matched_invoices = (
        db.query(Invoice)
        .filter(Invoice.invoice_no.in_(unique_nos))
        .all()
    )
    matched_nos = {inv.invoice_no for inv in matched_invoices}
    unmatched_nos = [no for no in unique_nos if no not in matched_nos]

    # 批量更新核销状态
    now = datetime.now(timezone.utc)
    for inv in matched_invoices:
        inv.is_reimbursed = True
        inv.reimbursed_at = now

    # 保存核销记录
    record = ReimbursementRecord(
        original_filename=filename,
        uploaded_by=current_user.id,
        total_count=len(unique_nos),
        matched_count=len(matched_invoices),
        unmatched_count=len(unmatched_nos),
        unmatched_details=json.dumps(unmatched_nos, ensure_ascii=False),
    )
    db.add(record)
    db.commit()

    logger.info(
        f"核销完成: 文件={filename}, 总数={len(unique_nos)}, "
        f"匹配={len(matched_invoices)}, 未匹配={len(unmatched_nos)}"
    )

    return ReimbursementResult(
        total_count=len(unique_nos),
        matched_count=len(matched_invoices),
        unmatched_count=len(unmatched_nos),
        unmatched_details=unmatched_nos,
    )


# ============================================================
# 核销历史记录
# ============================================================
@router.get("/records", response_model=dict)
async def list_records(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取核销历史记录（分页）。"""
    query = db.query(ReimbursementRecord)
    total = query.count()
    records = (
        query.order_by(ReimbursementRecord.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    items = []
    for r in records:
        # 获取上传者用户名
        uploader_username = None
        if r.uploaded_by:
            user = db.query(User).filter(User.id == r.uploaded_by).first()
            if user:
                uploader_username = user.username

        unmatched = []
        if r.unmatched_details:
            try:
                unmatched = json.loads(r.unmatched_details)
            except (json.JSONDecodeError, TypeError):
                unmatched = []

        items.append(
            ReimbursementRecordResponse(
                id=r.id,
                original_filename=r.original_filename,
                uploaded_by=r.uploaded_by,
                uploader_username=uploader_username,
                total_count=r.total_count,
                matched_count=r.matched_count,
                unmatched_count=r.unmatched_count,
                unmatched_details=unmatched,
                created_at=r.created_at,
            )
        )

    return {"total": total, "page": page, "page_size": page_size, "items": items}


# ============================================================
# 全量发票明细单 - 表头映射
# ============================================================
DETAIL_HEADER_MAP = {
    "序号": "serial_no",
    "发票代码": "invoice_code",
    "发票号码": "invoice_no",
    "数电发票号码": "digital_invoice_no",
    "销方识别号": "seller_tax_no",
    "销方名称": "seller_name",
    "购方识别号": "buyer_tax_no",
    "购买方名称": "buyer_name",
    "开票日期": "invoice_date",
    "金额": "amount",
    "税额": "tax_amount",
    "价税合计": "total_amount",
    "发票来源": "invoice_source",
    "发票票种": "invoice_type",
    "发票状态": "invoice_status",
    "是否正数发票": "is_positive",
    "发票风险等级": "risk_level",
    "开票人": "issuer",
    "备注": "remark",
}


def _cell_to_str(value) -> Optional[str]:
    """将 Excel 单元格值转为字符串（None 保持 None，空字符串返回 None）。"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    return s if s else None


def _is_empty_or_dash(value: Optional[str]) -> bool:
    """判断值是否为空或仅为横杠符号（用于检测汇总行）。"""
    if value is None:
        return True
    v = value.strip()
    return v == "" or v == "-" or v == "—" or v == "－"


def _build_detail_field_map(header: tuple) -> dict[int, str]:
    """根据表头构建“列索引 -> 模型字段名”映射。"""
    field_map: dict[int, str] = {}
    for idx, cell in enumerate(header):
        if cell is None:
            continue
        name = str(cell).strip()
        if name in DETAIL_HEADER_MAP:
            field_map[idx] = DETAIL_HEADER_MAP[name]
    return field_map


def _to_iso(dt) -> Optional[str]:
    """DateTime -> ISO 字符串。"""
    if dt is None:
        return None
    if isinstance(dt, datetime):
        return dt.isoformat()
    return str(dt)


def _detail_to_response(d: InvoiceDetail) -> InvoiceDetailResponse:
    return InvoiceDetailResponse(
        id=d.id,
        upload_batch_id=d.upload_batch_id,
        serial_no=d.serial_no,
        invoice_code=d.invoice_code,
        invoice_no=d.invoice_no,
        digital_invoice_no=d.digital_invoice_no,
        seller_tax_no=d.seller_tax_no,
        seller_name=d.seller_name,
        buyer_tax_no=d.buyer_tax_no,
        buyer_name=d.buyer_name,
        invoice_date=d.invoice_date,
        amount=d.amount,
        tax_amount=d.tax_amount,
        total_amount=d.total_amount,
        invoice_source=d.invoice_source,
        invoice_type=d.invoice_type,
        invoice_status=d.invoice_status,
        is_positive=d.is_positive,
        risk_level=d.risk_level,
        issuer=d.issuer,
        remark=d.remark,
        verify_status=d.verify_status,
        verified_at=_to_iso(d.verified_at),
        match_method=d.match_method,
        reimburse_status=d.reimburse_status,
        created_at=_to_iso(d.created_at),
    )


# ============================================================
# 上传全量发票明细单
# ============================================================
@router.post("/upload-detail", response_model=UploadDetailResult)
async def upload_detail(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """上传全量发票明细单（xlsx），解析并批量保存。"""
    filename = file.filename or "unknown"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext != "xlsx":
        raise HTTPException(status_code=400, detail="仅支持 xlsx 格式文件")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="文件内容为空")

    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)

        # 检查是否存在“发票基础信息”工作表
        target_sheet = "发票基础信息"
        if target_sheet not in wb.sheetnames:
            wb.close()
            raise HTTPException(
                status_code=400,
                detail=f"Excel文件中未找到'{target_sheet}'工作表，请上传符合要求的文件"
            )

        ws = wb[target_sheet]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析 Excel 失败: {e}")

    if not rows or len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 文件为空或仅含表头")

    header = rows[0]
    field_map = _build_detail_field_map(header)
    if not field_map:
        raise HTTPException(status_code=400, detail="未识别到有效表头列，请核对模板")

    # 验证表头是否包含所有必需列
    expected_columns = set(DETAIL_HEADER_MAP.keys())
    actual_columns = set()
    for cell in header:
        if cell is not None:
            actual_columns.add(str(cell).strip())

    missing_columns = expected_columns - actual_columns
    if missing_columns:
        raise HTTPException(
            status_code=400,
            detail=f"'发票基础信息'工作表缺少以下列：{', '.join(sorted(missing_columns))}，请上传符合要求的文件"
        )

    # 创建上传记录
    upload_log = InvoiceDetailUploadLog(
        original_filename=filename,
        total_count=0,
        uploaded_by=current_user.id,
    )
    db.add(upload_log)
    db.flush()  # 获取 id

    # 解析所有数据行（仅生成字段字典，先不构造 ORM 对象，便于去重）
    parsed_rows: list[dict] = []
    for row in rows[1:]:
        # 跳过空行
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        record: dict = {}
        for idx, field in field_map.items():
            if idx < len(row):
                record[field] = _cell_to_str(row[idx])

        # 跳过汇总行：数电发票号码和发票号码都为空或"-"的行视为汇总行
        digital_no = record.get("digital_invoice_no")
        inv_no = record.get("invoice_no")
        if _is_empty_or_dash(digital_no) and _is_empty_or_dash(inv_no):
            continue

        parsed_rows.append(record)

    if not parsed_rows:
        db.rollback()
        raise HTTPException(status_code=400, detail="文件中未解析到有效数据行")

    total_parsed = len(parsed_rows)

    # 基于 digital_invoice_no 去重：与数据库中已存在的记录做对比
    all_digital_nos = [
        row["digital_invoice_no"] for row in parsed_rows if row.get("digital_invoice_no")
    ]
    if all_digital_nos:
        existing = (
            db.query(InvoiceDetail.digital_invoice_no)
            .filter(InvoiceDetail.digital_invoice_no.in_(all_digital_nos))
            .all()
        )
        existing_nos = set(r[0] for r in existing)
    else:
        existing_nos = set()

    # 只保留 digital_invoice_no 为空（无法判断重复）或不在已存在集合中的记录
    new_rows = [
        row
        for row in parsed_rows
        if not row.get("digital_invoice_no")
        or row["digital_invoice_no"] not in existing_nos
    ]
    skipped_count = total_parsed - len(new_rows)

    # 构造 ORM 对象
    details: list[InvoiceDetail] = []
    for record in new_rows:
        kwargs = {"upload_batch_id": upload_log.id, "verify_status": "待核销"}
        kwargs.update(record)
        details.append(InvoiceDetail(**kwargs))

    if details:
        db.bulk_save_objects(details)

    # 实际插入条数（去重后）
    upload_log.total_count = len(details)
    db.commit()
    db.refresh(upload_log)

    logger.info(
        f"上传全量发票明细完成: 文件={filename}, batch_id={upload_log.id}, "
        f"解析={total_parsed}, 实际入库={len(details)}, 跳过重复={skipped_count}"
    )

    return UploadDetailResult(
        batch_id=upload_log.id,
        filename=filename,
        total_count=len(details),
        skipped_count=skipped_count,
    )


# ============================================================
# 执行核销
# ============================================================
@router.post("/verify", response_model=VerifyResult)
async def verify_details(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """对所有“待核销”的明细记录执行核销，与已拉取的邮件匹配。

    匹配规则：取 InvoiceDetail.digital_invoice_no，依次检查 ReimbEmailMessage.subject
    与 ReimbEmailMessage.attachment_name 是否包含该号码，命中任一即视为已核销。
    """
    pending_details = (
        db.query(InvoiceDetail)
        .filter(InvoiceDetail.verify_status == "待核销")
        .all()
    )
    total_count = len(pending_details)
    if total_count == 0:
        return VerifyResult(total_count=0, matched_count=0, unmatched_count=0, unmatched_details=[])

    # 拉取所有邮件记录，仅关心 subject 与 attachment_name
    email_messages = db.query(ReimbEmailMessage).all()
    subjects = [m.subject for m in email_messages if m.subject]
    attachment_names = [m.attachment_name for m in email_messages if m.attachment_name]

    now = datetime.now(timezone.utc)

    # 更新明细状态
    matched_count = 0
    unmatched_details: list[str] = []
    for d in pending_details:
        digital_no = (d.digital_invoice_no or "").strip()
        matched = False
        if digital_no:
            # a) 先在邮件主题中查找
            if any(digital_no in s for s in subjects):
                matched = True
            # b) 主题未命中再在附件名中查找
            elif any(digital_no in a for a in attachment_names):
                matched = True

        if matched:
            d.verify_status = "已核销"
            d.verified_at = now
            d.match_method = "邮箱匹配"
            matched_count += 1
        else:
            d.verify_status = "未匹配"
            ref = digital_no or (d.invoice_no or "").strip()
            if ref:
                unmatched_details.append(ref)

    unmatched_count = total_count - matched_count

    # 保存核销历史
    record = ReimbursementRecord(
        original_filename=f"核销_{now.strftime('%Y%m%d_%H%M%S')}",
        uploaded_by=current_user.id,
        total_count=total_count,
        matched_count=matched_count,
        unmatched_count=unmatched_count,
        unmatched_details=json.dumps(unmatched_details, ensure_ascii=False),
    )
    db.add(record)
    db.commit()

    logger.info(
        f"明细核销完成: 总数={total_count}, 匹配={matched_count}, 未匹配={unmatched_count}"
    )

    return VerifyResult(
        total_count=total_count,
        matched_count=matched_count,
        unmatched_count=unmatched_count,
        unmatched_details=unmatched_details,
    )


# ============================================================
# 分页查询全量发票明细
# ============================================================
@router.get("/details", response_model=InvoiceDetailListResponse)
async def list_details(
    keyword: Optional[str] = Query(None),
    invoice_source: Optional[str] = Query(None),
    invoice_type: Optional[str] = Query(None),
    verify_status: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    upload_batch_id: Optional[int] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """分页查询全量发票明细。"""
    query = db.query(InvoiceDetail)

    if keyword:
        kw = f"%{keyword}%"
        query = query.filter(
            or_(
                InvoiceDetail.seller_name.like(kw),
                InvoiceDetail.buyer_name.like(kw),
            )
        )
    if invoice_source:
        query = query.filter(InvoiceDetail.invoice_source == invoice_source)
    if invoice_type:
        query = query.filter(InvoiceDetail.invoice_type == invoice_type)
    if verify_status:
        query = query.filter(InvoiceDetail.verify_status == verify_status)
    if upload_batch_id is not None:
        query = query.filter(InvoiceDetail.upload_batch_id == upload_batch_id)
    if start_date:
        query = query.filter(InvoiceDetail.invoice_date >= start_date)
    if end_date:
        query = query.filter(InvoiceDetail.invoice_date <= end_date)

    total = query.count()
    records = (
        query.order_by(InvoiceDetail.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    items = [_detail_to_response(r) for r in records]
    return InvoiceDetailListResponse(
        total=total, page=page, page_size=page_size, items=items
    )


# ============================================================
# 获取上传历史记录
# ============================================================
@router.get("/upload-logs", response_model=UploadLogListResponse)
async def list_upload_logs(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取全量发票明细上传历史记录。"""
    query = db.query(InvoiceDetailUploadLog)
    total = query.count()
    records = (
        query.order_by(InvoiceDetailUploadLog.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    items: list[UploadLogResponse] = []
    for r in records:
        uploader_username = None
        if r.uploaded_by:
            user = db.query(User).filter(User.id == r.uploaded_by).first()
            if user:
                uploader_username = user.username
        items.append(
            UploadLogResponse(
                id=r.id,
                original_filename=r.original_filename,
                total_count=r.total_count,
                uploaded_by=r.uploaded_by,
                uploader_username=uploader_username,
                created_at=_to_iso(r.created_at),
            )
        )

    return UploadLogListResponse(
        total=total, page=page, page_size=page_size, items=items
    )


# ============================================================
# 报销单管理 - 独立邮箱配置
# ============================================================
def _reimb_config_to_response(config: ReimbEmailConfig) -> ReimbEmailConfigResponse:
    return ReimbEmailConfigResponse(
        id=config.id,
        email_address=config.email_address,
        imap_server=config.imap_server,
        port=config.port,
        use_ssl=config.use_ssl,
        is_active=config.is_active,
        last_check_at=config.last_check_at.isoformat() if config.last_check_at else None,
        created_at=config.created_at.isoformat() if config.created_at else None,
    )


@router.post("/email-configs", response_model=ReimbEmailConfigResponse)
async def create_reimb_email_config(
    data: ReimbEmailConfigCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """添加报销单管理专用邮箱配置。"""
    config = ReimbEmailConfig(
        email_address=data.email_address,
        imap_server=data.imap_server,
        port=data.port,
        password_encrypted=encrypt_password(data.password),
        use_ssl=data.use_ssl,
        user_id=current_user.id,
    )
    db.add(config)
    db.commit()
    db.refresh(config)
    return _reimb_config_to_response(config)


@router.get("/email-configs", response_model=list[ReimbEmailConfigResponse])
async def list_reimb_email_configs(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取报销单管理邮箱配置列表。"""
    configs = db.query(ReimbEmailConfig).all()
    return [_reimb_config_to_response(c) for c in configs]


@router.put("/email-configs/{config_id}", response_model=ReimbEmailConfigResponse)
async def update_reimb_email_config(
    config_id: int,
    data: ReimbEmailConfigUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """更新报销单管理邮箱配置。"""
    config = db.query(ReimbEmailConfig).filter(ReimbEmailConfig.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail="邮箱配置不存在")
    if data.email_address is not None:
        config.email_address = data.email_address
    if data.imap_server is not None:
        config.imap_server = data.imap_server
    if data.port is not None:
        config.port = data.port
    if data.use_ssl is not None:
        config.use_ssl = data.use_ssl
    if data.is_active is not None:
        config.is_active = data.is_active
    if data.password:
        config.password_encrypted = encrypt_password(data.password)
    db.commit()
    db.refresh(config)
    return _reimb_config_to_response(config)


@router.delete("/email-configs/{config_id}")
async def delete_reimb_email_config(
    config_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """删除报销单管理邮箱配置。"""
    config = db.query(ReimbEmailConfig).filter(ReimbEmailConfig.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail="邮箱配置不存在")
    # 级联删除相关日志和消息
    db.query(ReimbEmailFetchLog).filter(ReimbEmailFetchLog.config_id == config_id).delete()
    db.query(ReimbEmailMessage).filter(ReimbEmailMessage.config_id == config_id).delete()
    db.delete(config)
    db.commit()
    return {"detail": "已删除"}


@router.post("/email-configs/{config_id}/test", response_model=ReimbEmailTestResult)
async def test_reimb_email_connection(
    config_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """测试邮箱连接。"""
    config = db.query(ReimbEmailConfig).filter(ReimbEmailConfig.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail="邮箱配置不存在")
    try:
        password = decrypt_password(config.password_encrypted)
        if config.use_ssl:
            mail = imaplib.IMAP4_SSL(config.imap_server, config.port)
        else:
            mail = imaplib.IMAP4(config.imap_server, config.port)
        mail.login(config.email_address, password)
        mail.logout()
        return ReimbEmailTestResult(success=True, message="连接成功")
    except Exception as e:
        # 清理IMAP错误中的bytes表示
        err_msg = str(e)
        if e.args and isinstance(e.args[0], bytes):
            err_msg = e.args[0].decode("utf-8", errors="replace")
        elif e.args and len(e.args) > 1 and isinstance(e.args[1], bytes):
            err_msg = e.args[1].decode("utf-8", errors="replace")
        return ReimbEmailTestResult(success=False, message=f"连接失败: {err_msg}")


@router.post("/email-fetch/{config_id}")
async def fetch_reimb_emails(
    config_id: int,
    req: ReimbEmailFetchRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """从配置邮箱拉取邮件（同步执行，基于日期范围）。"""
    config = db.query(ReimbEmailConfig).filter(ReimbEmailConfig.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail="邮箱配置不存在")

    try:
        password = decrypt_password(config.password_encrypted)

        # 建立IMAP连接
        if config.use_ssl:
            mail = imaplib.IMAP4_SSL(config.imap_server, config.port)
        else:
            mail = imaplib.IMAP4(config.imap_server, config.port)
        mail.login(config.email_address, password)
        mail.select("INBOX")

        # 构造搜索条件
        search_criteria = []
        if req.date_from:
            d = datetime.strptime(req.date_from, "%Y-%m-%d")
            search_criteria.append(f'SINCE {d.strftime("%d-%b-%Y")}')
        if req.date_to:
            d = datetime.strptime(req.date_to, "%Y-%m-%d")
            # IMAP BEFORE是不含当天的，所以加1天
            d2 = d + timedelta(days=1)
            search_criteria.append(f'BEFORE {d2.strftime("%d-%b-%Y")}')

        criteria_str = " ".join(search_criteria) if search_criteria else "ALL"

        status, data = mail.uid("SEARCH", None, criteria_str)
        if status != "OK":
            raise Exception("搜索邮件失败")

        uids = data[0].split() if data[0] else []
        total_checked = len(uids)

        # 去重：检查已存在的UID
        existing_uids = set()
        if uids:
            uid_strs = [u.decode() for u in uids]
            existing = (
                db.query(ReimbEmailMessage.message_uid)
                .filter(ReimbEmailMessage.config_id == config_id)
                .filter(ReimbEmailMessage.message_uid.in_(uid_strs))
                .all()
            )
            existing_uids = {r[0] for r in existing}

        new_uids = [u for u in uids if u.decode() not in existing_uids]
        new_count = 0

        # 确保存储目录
        today = datetime.now().strftime("%Y-%m-%d")
        target_dir = os.path.join(settings.REIMB_EMAIL_TEMP_PATH, today)
        os.makedirs(target_dir, exist_ok=True)

        for uid in new_uids:
            uid_str = uid.decode()
            try:
                fstatus, msg_data = mail.uid("FETCH", uid, "(RFC822)")
                if fstatus != "OK" or not msg_data or not msg_data[0]:
                    continue
                raw = msg_data[0][1]
                msg = email_lib.message_from_bytes(raw)

                # 解码主题
                subject = ""
                raw_subject = msg.get("Subject", "")
                if raw_subject:
                    decoded_parts = decode_header(raw_subject)
                    parts = []
                    for part, charset in decoded_parts:
                        if isinstance(part, bytes):
                            parts.append(part.decode(charset or "utf-8", errors="replace"))
                        else:
                            parts.append(part)
                    subject = "".join(parts)

                # 发件人
                sender = msg.get("From", "")

                # 邮件时间
                received_at = None
                date_str = msg.get("Date")
                if date_str:
                    try:
                        received_at = parsedate_to_datetime(date_str)
                    except Exception:
                        pass

                # 提取附件
                has_attachment = False
                if msg.is_multipart():
                    for part in msg.walk():
                        content_disposition = part.get("Content-Disposition", "")
                        if "attachment" not in content_disposition:
                            continue
                        raw_fname = part.get_filename()
                        if not raw_fname:
                            continue
                        # 解码文件名
                        decoded_fname_parts = decode_header(raw_fname)
                        fname_parts = []
                        for fp, fcharset in decoded_fname_parts:
                            if isinstance(fp, bytes):
                                fname_parts.append(fp.decode(fcharset or "utf-8", errors="replace"))
                            else:
                                fname_parts.append(fp)
                        filename = "".join(fname_parts).strip()
                        if not filename:
                            continue

                        # 只保存PDF文件
                        if not filename.lower().endswith(".pdf"):
                            continue

                        # 保存附件
                        file_data = part.get_payload(decode=True)
                        if not file_data:
                            continue

                        # 安全文件名
                        safe_name = filename.replace("/", "_").replace("\\", "_")
                        if len(safe_name) > 200:
                            safe_name = safe_name[:200]
                        file_path = os.path.join(target_dir, safe_name)
                        # 避免重名
                        counter = 1
                        base, ext = os.path.splitext(file_path)
                        while os.path.exists(file_path):
                            file_path = f"{base}_{counter}{ext}"
                            counter += 1

                        with open(file_path, "wb") as f:
                            f.write(file_data)

                        # 创建消息记录
                        email_msg = ReimbEmailMessage(
                            config_id=config_id,
                            message_uid=uid_str,
                            subject=subject[:500] if subject else None,
                            sender=sender[:200] if sender else None,
                            received_at=received_at,
                            attachment_name=filename[:300],
                            attachment_path=file_path,
                            file_size=len(file_data),
                            user_id=current_user.id,
                        )
                        db.add(email_msg)
                        has_attachment = True
                        new_count += 1

                # 如果没有附件但有主题，也保存一条记录（用于主题匹配）
                if not has_attachment and subject:
                    email_msg = ReimbEmailMessage(
                        config_id=config_id,
                        message_uid=uid_str,
                        subject=subject[:500],
                        sender=sender[:200] if sender else None,
                        received_at=received_at,
                        attachment_name=None,
                        attachment_path=None,
                        file_size=None,
                        user_id=current_user.id,
                    )
                    db.add(email_msg)
                    new_count += 1

            except Exception as e:
                logger.warning(f"处理邮件 UID={uid_str} 失败: {e}")
                continue

        mail.logout()

        # 更新最后检查时间
        config.last_check_at = datetime.now(timezone.utc)

        # 保存拉取日志
        fetch_log = ReimbEmailFetchLog(
            config_id=config_id,
            email_address=config.email_address,
            total_emails_checked=total_checked,
            new_emails_count=new_count,
            status="success",
        )
        db.add(fetch_log)
        db.commit()

        return {
            "success": True,
            "total_checked": total_checked,
            "new_count": new_count,
            "skipped": total_checked - len(new_uids),
        }

    except HTTPException:
        raise
    except Exception as e:
        # 记录失败日志
        fetch_log = ReimbEmailFetchLog(
            config_id=config_id,
            email_address=config.email_address,
            total_emails_checked=0,
            new_emails_count=0,
            status="failed",
            error_message=str(e)[:1000],
        )
        db.add(fetch_log)
        db.commit()
        raise HTTPException(status_code=500, detail=f"邮件拉取失败: {str(e)}")


@router.get("/email-messages")
async def list_reimb_email_messages(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    config_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取报销单管理拉取的邮件列表。"""
    query = db.query(ReimbEmailMessage)
    if config_id is not None:
        query = query.filter(ReimbEmailMessage.config_id == config_id)
    total = query.count()
    records = (
        query.order_by(ReimbEmailMessage.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    items = []
    for r in records:
        items.append({
            "id": r.id,
            "config_id": r.config_id,
            "message_uid": r.message_uid,
            "subject": r.subject,
            "sender": r.sender,
            "received_at": r.received_at.isoformat() if r.received_at else None,
            "attachment_name": r.attachment_name,
            "file_size": r.file_size,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        })
    return {"total": total, "page": page, "page_size": page_size, "items": items}


@router.get("/email-fetch-logs")
async def list_reimb_email_fetch_logs(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取报销单管理邮件拉取日志。"""
    query = db.query(ReimbEmailFetchLog)
    total = query.count()
    records = (
        query.order_by(ReimbEmailFetchLog.fetch_time.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    items = []
    for r in records:
        items.append({
            "id": r.id,
            "config_id": r.config_id,
            "email_address": r.email_address,
            "fetch_time": r.fetch_time.isoformat() if r.fetch_time else None,
            "total_emails_checked": r.total_emails_checked,
            "new_emails_count": r.new_emails_count,
            "status": r.status,
            "error_message": r.error_message,
        })
    return {"total": total, "page": page, "page_size": page_size, "items": items}


# ============================================================
# 手工匹配
# ============================================================
MANUAL_MATCH_ALLOWED_EXTS = (".pdf", ".jpg", ".jpeg", ".png")


def _pdf_to_image_temp(file_path: str) -> str:
    """将 PDF 首页转为 PNG 临时文件，返回临时路径。"""
    if fitz is None:
        raise RuntimeError("未安装 PyMuPDF (fitz)，无法处理 PDF。请执行 pip install PyMuPDF")
    doc = fitz.open(file_path)
    try:
        if doc.page_count < 1:
            raise ValueError(f"PDF 无可用页面: {file_path}")
        page = doc[0]
        mat = fitz.Matrix(2.0, 2.0)
        pix = page.get_pixmap(matrix=mat, alpha=False)
        fd, temp_path = tempfile.mkstemp(suffix=".png", prefix="manual_match_")
        os.close(fd)
        pix.save(temp_path)
    finally:
        doc.close()
    return temp_path


def _image_file_to_base64(file_path: str) -> str:
    with open(file_path, "rb") as f:
        data = f.read()
    return base64.b64encode(data).decode("utf-8")


def _extract_json_from_text(text: str) -> Optional[dict]:
    """从模型返回文本中提取 JSON。"""
    if not text:
        return None
    json_match = re.search(r"```(?:json)?\s*\n?(.*?)\n?```", text, re.DOTALL)
    if json_match:
        candidate = json_match.group(1).strip()
    else:
        candidate = text.strip()
    try:
        return json.loads(candidate)
    except json.JSONDecodeError:
        start = text.find("{")
        end = text.rfind("}")
        if start != -1 and end != -1 and end > start:
            try:
                return json.loads(text[start:end + 1])
            except json.JSONDecodeError:
                return None
        return None


async def _recognize_digital_invoice_no(file_path: str) -> dict:
    """调用视觉模型识别数电发票号码。返回 {success, invoice_no, raw, error}。"""
    api_key = settings.DASHSCOPE_API_KEY
    if not api_key or api_key == "your-dashscope-api-key":
        return {"success": False, "invoice_no": None, "raw": None, "error": "DASHSCOPE_API_KEY未配置"}

    ext = Path(file_path).suffix.lower()
    temp_image_path: Optional[str] = None
    try:
        if ext == ".pdf":
            temp_image_path = _pdf_to_image_temp(file_path)
            img_base64 = _image_file_to_base64(temp_image_path)
            mime_type = "image/png"
        elif ext in (".jpg", ".jpeg"):
            img_base64 = _image_file_to_base64(file_path)
            mime_type = "image/jpeg"
        elif ext == ".png":
            img_base64 = _image_file_to_base64(file_path)
            mime_type = "image/png"
        else:
            return {"success": False, "invoice_no": None, "raw": None, "error": f"不支持的文件格式: {ext}"}
    except Exception as e:
        logger.error(f"手工匹配 - 文件预处理失败: {file_path}, {e}")
        return {"success": False, "invoice_no": None, "raw": None, "error": f"文件处理失败: {e}"}
    finally:
        if temp_image_path and os.path.exists(temp_image_path):
            try:
                os.remove(temp_image_path)
            except OSError:
                pass

    prompt = "请识别这张发票图片中的数电发票号码（20位数字）。仅返回JSON格式：{\"digital_invoice_no\": \"号码\"}"
    api_url = f"{settings.DASHSCOPE_BASE_URL.rstrip('/')}/chat/completions"
    try:
        async with httpx.AsyncClient(timeout=120.0) as client:
            response = await client.post(
                api_url,
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": settings.VISION_MODEL,
                    "messages": [
                        {
                            "role": "user",
                            "content": [
                                {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{img_base64}"}},
                                {"type": "text", "text": prompt},
                            ],
                        }
                    ],
                },
            )
        if response.status_code != 200:
            return {
                "success": False,
                "invoice_no": None,
                "raw": None,
                "error": f"视觉模型调用失败，状态码：{response.status_code}, 响应：{response.text[:200]}",
            }
        data = response.json()
        content = data["choices"][0]["message"]["content"]
    except httpx.TimeoutException:
        return {"success": False, "invoice_no": None, "raw": None, "error": "视觉模型调用超时"}
    except Exception as e:
        logger.error(f"手工匹配 - 识别调用异常: {e}", exc_info=True)
        return {"success": False, "invoice_no": None, "raw": None, "error": f"识别调用异常: {e}"}

    parsed = _extract_json_from_text(content)
    if not parsed or not isinstance(parsed, dict):
        return {
            "success": False,
            "invoice_no": None,
            "raw": {"content": content},
            "error": "未能从模型返回中解析出 JSON",
        }

    no = parsed.get("digital_invoice_no")
    no_str = str(no).strip() if no is not None else None
    if not no_str:
        return {"success": False, "invoice_no": None, "raw": parsed, "error": "未识别到数电发票号码"}

    return {"success": True, "invoice_no": no_str, "raw": parsed, "error": None}


@router.post("/manual-match/{detail_id}", response_model=ManualMatchRecognizeResult)
async def manual_match_invoice(
    detail_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """为未匹配的发票明细上传发票文件，调用视觉模型识别发票号码并手工核销。"""
    detail = db.query(InvoiceDetail).filter(InvoiceDetail.id == detail_id).first()
    if not detail:
        raise HTTPException(status_code=404, detail="发票明细不存在")
    if detail.verify_status != "未匹配":
        raise HTTPException(status_code=400, detail="仅「未匹配」状态的明细可进行手工匹配")

    filename = file.filename or "unknown"
    ext = Path(filename).suffix.lower()
    if ext not in MANUAL_MATCH_ALLOWED_EXTS:
        raise HTTPException(
            status_code=400,
            detail=f"仅支持以下格式：{', '.join(MANUAL_MATCH_ALLOWED_EXTS)}",
        )

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="文件内容为空")

    today = datetime.now().strftime("%Y-%m-%d")
    target_dir = os.path.join(settings.MANUAL_MATCH_PATH, today)
    os.makedirs(target_dir, exist_ok=True)

    safe_name = filename.replace("/", "_").replace("\\", "_")
    if len(safe_name) > 200:
        safe_name = safe_name[:200]
    file_path = os.path.join(target_dir, safe_name)
    base, ext_part = os.path.splitext(file_path)
    counter = 1
    while os.path.exists(file_path):
        file_path = f"{base}_{counter}{ext_part}"
        counter += 1
    with open(file_path, "wb") as f:
        f.write(content)

    recognition = await _recognize_digital_invoice_no(file_path)
    recognized_no = recognition.get("invoice_no")
    raw_data = recognition.get("raw")

    if not recognition.get("success"):
        record = ManualMatchRecord(
            invoice_detail_id=detail.id,
            original_filename=filename,
            file_path=file_path,
            recognized_invoice_no=None,
            recognized_data=json.dumps(raw_data, ensure_ascii=False) if raw_data else None,
            match_status="失败",
            operated_by=current_user.id,
        )
        db.add(record)
        db.commit()
        return ManualMatchRecognizeResult(
            success=False,
            recognized_invoice_no=None,
            recognized_data=raw_data if isinstance(raw_data, dict) else None,
            match_result="recognition_failed",
            message=recognition.get("error") or "识别失败",
        )

    target_no = (detail.digital_invoice_no or "").strip()
    is_matched = bool(target_no) and recognized_no == target_no

    if is_matched:
        detail.verify_status = "已核销"
        detail.match_method = "手工匹配"
        detail.verified_at = datetime.now(timezone.utc)
        match_status_text = "成功"
        match_result = "matched"
        message = "识别成功，发票号码与明细匹配，已核销"
    else:
        match_status_text = "失败"
        match_result = "not_matched"
        message = (
            f"识别出的发票号码（{recognized_no}）与明细发票号码（{target_no or '空'}）不一致"
        )

    record = ManualMatchRecord(
        invoice_detail_id=detail.id,
        original_filename=filename,
        file_path=file_path,
        recognized_invoice_no=recognized_no,
        recognized_data=json.dumps(raw_data, ensure_ascii=False) if raw_data else None,
        match_status=match_status_text,
        operated_by=current_user.id,
    )
    db.add(record)
    db.commit()

    return ManualMatchRecognizeResult(
        success=is_matched,
        recognized_invoice_no=recognized_no,
        recognized_data=raw_data if isinstance(raw_data, dict) else None,
        match_result=match_result,
        message=message,
    )


@router.post("/submit-reimburse")
async def submit_reimburse(
    payload: dict,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """批量提交报销 - 将选中的已核销发票标记为已报销"""
    ids = payload.get("ids", [])
    if not ids:
        raise HTTPException(status_code=400, detail="请选择要提交报销的发票")

    # 只允许对已核销的发票进行报销提交
    updated = db.query(InvoiceDetail).filter(
        InvoiceDetail.id.in_(ids),
        InvoiceDetail.verify_status == "已核销"
    ).update(
        {"reimburse_status": "已报销"},
        synchronize_session=False
    )
    db.commit()

    return {"success": True, "updated_count": updated, "total_requested": len(ids)}


@router.post("/create-reimburse-application")
async def create_reimburse_application(
    payload: ReimbursementApplicationCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """创建报销申请"""
    # 验证发票存在且已核销
    invoices = db.query(InvoiceDetail).filter(
        InvoiceDetail.id.in_(payload.invoice_ids),
        InvoiceDetail.verify_status == "已核销"
    ).all()

    if not invoices:
        raise HTTPException(status_code=400, detail="未找到有效的已核销发票")

    # 计算总金额
    total = 0.0
    for inv in invoices:
        try:
            total += float(inv.total_amount) if inv.total_amount else 0.0
        except (ValueError, TypeError):
            pass

    # 生成报销单编号
    today = datetime.now().strftime("%Y%m%d")
    today_count = db.query(ReimbursementApplication).filter(
        ReimbursementApplication.reimburse_no.like(f"RBS-{today}-%")
    ).count()
    reimburse_no = f"RBS-{today}-{str(today_count + 1).zfill(3)}"

    # 创建报销申请
    application = ReimbursementApplication(
        reimburse_no=reimburse_no,
        applicant_name=current_user.full_name or current_user.username,
        applicant_position=current_user.position,
        reimburse_date=payload.reimburse_date,
        department=payload.department,
        category=payload.category,
        reason=payload.reason,
        remark=payload.remark,
        total_amount=total,
        invoice_ids=json.dumps(payload.invoice_ids),
        status="已提交",
        submitted_by=current_user.id,
    )
    db.add(application)

    # 更新发票的报销状态
    for inv in invoices:
        inv.reimburse_status = "已报销"

    db.commit()
    db.refresh(application)

    return {
        "id": application.id,
        "reimburse_no": application.reimburse_no,
        "applicant_name": application.applicant_name,
        "applicant_position": application.applicant_position,
        "reimburse_date": application.reimburse_date,
        "department": application.department,
        "category": application.category,
        "reason": application.reason,
        "remark": application.remark,
        "total_amount": application.total_amount,
        "invoice_ids": payload.invoice_ids,
        "status": application.status,
        "created_at": application.created_at.isoformat() if application.created_at else None,
    }


@router.get("/manual-match-records")
async def list_manual_match_records(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """手工匹配记录分页查询。"""
    query = db.query(ManualMatchRecord)
    total = query.count()
    records = (
        query.order_by(ManualMatchRecord.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    items = []
    for r in records:
        items.append(ManualMatchRecordResponse(
            id=r.id,
            invoice_detail_id=r.invoice_detail_id,
            original_filename=r.original_filename,
            recognized_invoice_no=r.recognized_invoice_no,
            match_status=r.match_status,
            operated_by=r.operated_by,
            created_at=_to_iso(r.created_at),
        ))
    return {"total": total, "page": page, "page_size": page_size, "items": items}
