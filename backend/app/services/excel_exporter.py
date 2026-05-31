"""
Excel 导出服务

使用 openpyxl 生成发票汇总/明细 Excel 文件，支持三种模式：
1. monthly_summary - 月度汇总(按日期)
2. category_summary - 分类汇总
3. detail - 明细表

输出统一保存到 storage/exports/，文件名格式:
    invoices_{mode}_{year}_{month}_{timestamp}.xlsx
"""
from __future__ import annotations

import logging
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.models.invoice import Invoice

logger = logging.getLogger(__name__)


# 输出目录(相对于项目根目录)
EXPORT_OUTPUT_DIR = Path("storage/exports")


# ---------------------------------------------------------------------------
# 公共样式
# ---------------------------------------------------------------------------
TITLE_FONT = Font(name="微软雅黑", size=16, bold=True, color="1F3864")
SUBTITLE_FONT = Font(name="微软雅黑", size=10, color="606266")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="微软雅黑", size=10, color="303133")
TOTAL_FONT = Font(name="微软雅黑", size=11, bold=True, color="1F3864")

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
ROW_FILL_EVEN = PatternFill("solid", fgColor="FFFFFF")
ROW_FILL_ODD = PatternFill("solid", fgColor="F2F6FC")
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")

THIN_SIDE = Side(style="thin", color="B9C5D8")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
TOP_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=Side(style="medium", color="1F3864"),
    bottom=THIN_SIDE,
)

MONEY_FORMAT = "#,##0.00;[Red]-#,##0.00;-"


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------
def ensure_export_dir() -> Path:
    """确保导出目录存在并返回绝对路径"""
    out_dir = (Path.cwd() / EXPORT_OUTPUT_DIR).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir


def _build_filename(mode: str, year: int, month: int) -> Path:
    out_dir = ensure_export_dir()
    ts = int(time.time() * 1000)
    return out_dir / f"invoices_{mode}_{year}_{month:02d}_{ts}.xlsx"


def _safe_float(v) -> float:
    try:
        if v is None:
            return 0.0
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _parse_invoice_date(raw: Optional[str]) -> Optional[datetime]:
    """尽量将发票日期字符串解析为 datetime"""
    if not raw:
        return None
    raw = raw.strip()
    if not raw:
        return None
    fmts = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%Y年%m月%d日",
        "%Y%m%d",
        "%Y-%m-%d %H:%M:%S",
    ]
    for fmt in fmts:
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def _filter_invoices(
    db: Session,
    year: int,
    month: int,
    category: Optional[str] = None,
    source_type: Optional[str] = None,
) -> List[Invoice]:
    """按条件加载发票，并按 invoice_date 升序排序。

    invoice_date 是字符串字段，先在数据库做 LIKE 粗筛，再在内存按解析后的日期精筛。
    """
    q = db.query(Invoice)
    if category:
        q = q.filter(Invoice.category == category)
    if source_type:
        q = q.filter(Invoice.source_type == source_type)

    rows: List[Invoice] = q.all()

    target_year = year
    target_month = month
    result: List[Tuple[Optional[datetime], Invoice]] = []
    for inv in rows:
        d = _parse_invoice_date(inv.invoice_date)
        if d is None:
            # 日期无法解析的发票跳过(避免错入其它月份)
            continue
        if d.year == target_year and d.month == target_month:
            result.append((d, inv))

    result.sort(key=lambda x: (x[0] or datetime.max, x[1].id))
    return [inv for _, inv in result]


def _apply_header(ws: Worksheet, headers: List[str], row_idx: int) -> None:
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[row_idx].height = 26


def _apply_title(ws: Worksheet, title: str, col_count: int, row_idx: int) -> None:
    ws.merge_cells(
        start_row=row_idx,
        start_column=1,
        end_row=row_idx,
        end_column=col_count,
    )
    cell = ws.cell(row=row_idx, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = CENTER_ALIGN
    ws.row_dimensions[row_idx].height = 32


def _apply_subtitle(ws: Worksheet, text: str, col_count: int, row_idx: int) -> None:
    ws.merge_cells(
        start_row=row_idx,
        start_column=1,
        end_row=row_idx,
        end_column=col_count,
    )
    cell = ws.cell(row=row_idx, column=1, value=text)
    cell.font = SUBTITLE_FONT
    cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row_idx].height = 18


def _autosize_columns(ws: Worksheet, col_count: int, min_width: int = 10) -> None:
    """根据每列内容长度估算列宽(中文按2宽度计算)"""
    for col_idx in range(1, col_count + 1):
        letter = get_column_letter(col_idx)
        max_len = min_width
        for cell in ws[letter]:
            value = cell.value
            if value is None:
                continue
            text = str(value)
            length = 0
            for ch in text:
                length += 2 if ord(ch) > 127 else 1
            if length > max_len:
                max_len = length
        ws.column_dimensions[letter].width = min(max_len + 4, 60)


def _apply_data_row(
    ws: Worksheet,
    row_idx: int,
    values: List,
    money_cols: List[int],
    is_odd: bool = False,
) -> None:
    fill = ROW_FILL_ODD if is_odd else ROW_FILL_EVEN
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = DATA_FONT
        cell.fill = fill
        cell.border = THIN_BORDER
        if col_idx in money_cols:
            cell.alignment = RIGHT_ALIGN
            cell.number_format = MONEY_FORMAT
        elif isinstance(value, (int, float)) and not isinstance(value, bool):
            cell.alignment = RIGHT_ALIGN
        else:
            cell.alignment = LEFT_ALIGN
    ws.row_dimensions[row_idx].height = 20


def _apply_total_row(
    ws: Worksheet,
    row_idx: int,
    values: List,
    money_cols: List[int],
) -> None:
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = TOP_BORDER
        if col_idx in money_cols:
            cell.alignment = RIGHT_ALIGN
            cell.number_format = MONEY_FORMAT
        elif col_idx == 1:
            cell.alignment = CENTER_ALIGN
        elif isinstance(value, (int, float)) and not isinstance(value, bool):
            cell.alignment = RIGHT_ALIGN
        else:
            cell.alignment = LEFT_ALIGN
    ws.row_dimensions[row_idx].height = 24


# ---------------------------------------------------------------------------
# 主导出器
# ---------------------------------------------------------------------------
class ExcelExporter:
    """Excel 导出服务"""

    # ------------------------------------------------------------------ #
    # 月度汇总
    # ------------------------------------------------------------------ #
    def export_monthly_summary(
        self,
        db: Session,
        year: int,
        month: int,
        category: Optional[str] = None,
        source_type: Optional[str] = None,
    ) -> Tuple[str, int]:
        """月度汇总表：按日期汇总每天的发票数量与金额。

        :return: (输出文件绝对路径, 涉及发票数量)
        """
        invoices = _filter_invoices(db, year, month, category, source_type)
        wb = Workbook()
        ws = wb.active
        ws.title = "月度汇总"

        headers = ["日期", "发票数量", "不含税总额", "税额合计", "价税合计"]
        col_count = len(headers)
        money_cols = [3, 4, 5]

        # 标题
        _apply_title(ws, f"{year}年{month:02d}月发票汇总表", col_count, 1)
        _apply_subtitle(
            ws,
            f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}　共 {len(invoices)} 张",
            col_count,
            2,
        )
        _apply_header(ws, headers, 3)

        # 按日期分组
        groups: Dict[str, Dict[str, float]] = {}
        for inv in invoices:
            d = _parse_invoice_date(inv.invoice_date)
            key = d.strftime("%Y-%m-%d") if d else "未知日期"
            grp = groups.setdefault(
                key, {"count": 0, "amount": 0.0, "tax": 0.0, "total": 0.0}
            )
            grp["count"] += 1
            grp["amount"] += _safe_float(inv.amount)
            grp["tax"] += _safe_float(inv.tax)
            grp["total"] += _safe_float(inv.total)

        sorted_keys = sorted(groups.keys())
        total_count = 0
        total_amount = 0.0
        total_tax = 0.0
        total_total = 0.0

        row_idx = 4
        for i, key in enumerate(sorted_keys):
            g = groups[key]
            _apply_data_row(
                ws,
                row_idx,
                [key, int(g["count"]), g["amount"], g["tax"], g["total"]],
                money_cols=money_cols,
                is_odd=(i % 2 == 1),
            )
            total_count += int(g["count"])
            total_amount += g["amount"]
            total_tax += g["tax"]
            total_total += g["total"]
            row_idx += 1

        if not sorted_keys:
            cell = ws.cell(row=row_idx, column=1, value="本月暂无发票数据")
            ws.merge_cells(
                start_row=row_idx,
                start_column=1,
                end_row=row_idx,
                end_column=col_count,
            )
            cell.font = SUBTITLE_FONT
            cell.alignment = CENTER_ALIGN
            row_idx += 1

        # 合计
        _apply_total_row(
            ws,
            row_idx,
            ["合计", total_count, total_amount, total_tax, total_total],
            money_cols=money_cols,
        )

        ws.freeze_panes = "A4"
        _autosize_columns(ws, col_count)

        out_path = _build_filename("monthly", year, month)
        wb.save(out_path)
        logger.info(f"生成月度汇总Excel: {out_path} (发票数 {len(invoices)})")
        return str(out_path), len(invoices)

    # ------------------------------------------------------------------ #
    # 分类汇总
    # ------------------------------------------------------------------ #
    def export_category_summary(
        self,
        db: Session,
        year: int,
        month: int,
        category: Optional[str] = None,
        source_type: Optional[str] = None,
    ) -> Tuple[str, int]:
        """分类汇总表：按分类汇总。"""
        invoices = _filter_invoices(db, year, month, category, source_type)
        wb = Workbook()
        ws = wb.active
        ws.title = "分类汇总"

        headers = ["分类", "发票数量", "不含税总额", "税额合计", "价税合计", "金额占比"]
        col_count = len(headers)
        money_cols = [3, 4, 5]

        _apply_title(ws, f"{year}年{month:02d}月发票分类汇总表", col_count, 1)
        _apply_subtitle(
            ws,
            f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}　共 {len(invoices)} 张",
            col_count,
            2,
        )
        _apply_header(ws, headers, 3)

        groups: Dict[str, Dict[str, float]] = {}
        for inv in invoices:
            key = inv.category or "未分类"
            grp = groups.setdefault(
                key, {"count": 0, "amount": 0.0, "tax": 0.0, "total": 0.0}
            )
            grp["count"] += 1
            grp["amount"] += _safe_float(inv.amount)
            grp["tax"] += _safe_float(inv.tax)
            grp["total"] += _safe_float(inv.total)

        # 按价税合计降序
        sorted_items = sorted(
            groups.items(), key=lambda kv: kv[1]["total"], reverse=True
        )
        total_total = sum(g["total"] for g in groups.values()) or 0.0

        total_count = 0
        sum_amount = 0.0
        sum_tax = 0.0
        sum_total = 0.0

        row_idx = 4
        for i, (cat, g) in enumerate(sorted_items):
            ratio = (g["total"] / total_total) if total_total > 0 else 0.0
            ratio_text = f"{ratio * 100:.2f}%"
            _apply_data_row(
                ws,
                row_idx,
                [
                    cat,
                    int(g["count"]),
                    g["amount"],
                    g["tax"],
                    g["total"],
                    ratio_text,
                ],
                money_cols=money_cols,
                is_odd=(i % 2 == 1),
            )
            total_count += int(g["count"])
            sum_amount += g["amount"]
            sum_tax += g["tax"]
            sum_total += g["total"]
            row_idx += 1

        if not sorted_items:
            cell = ws.cell(row=row_idx, column=1, value="本月暂无发票数据")
            ws.merge_cells(
                start_row=row_idx,
                start_column=1,
                end_row=row_idx,
                end_column=col_count,
            )
            cell.font = SUBTITLE_FONT
            cell.alignment = CENTER_ALIGN
            row_idx += 1

        _apply_total_row(
            ws,
            row_idx,
            ["合计", total_count, sum_amount, sum_tax, sum_total, "100.00%"],
            money_cols=money_cols,
        )

        ws.freeze_panes = "A4"
        _autosize_columns(ws, col_count)

        out_path = _build_filename("category", year, month)
        wb.save(out_path)
        logger.info(f"生成分类汇总Excel: {out_path} (发票数 {len(invoices)})")
        return str(out_path), len(invoices)

    # ------------------------------------------------------------------ #
    # 明细表
    # ------------------------------------------------------------------ #
    def export_detail(
        self,
        db: Session,
        year: int,
        month: int,
        category: Optional[str] = None,
        source_type: Optional[str] = None,
    ) -> Tuple[str, int]:
        """明细表：所有发票详细信息。"""
        invoices = _filter_invoices(db, year, month, category, source_type)
        wb = Workbook()
        ws = wb.active
        ws.title = "发票明细"

        headers = [
            "序号",
            "发票号码",
            "开票日期",
            "销售方",
            "购买方",
            "开票内容",
            "金额",
            "税额",
            "价税合计",
            "分类",
            "来源类型",
        ]
        col_count = len(headers)
        money_cols = [7, 8, 9]

        _apply_title(ws, f"{year}年{month:02d}月发票明细表", col_count, 1)
        _apply_subtitle(
            ws,
            f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}　共 {len(invoices)} 张",
            col_count,
            2,
        )
        _apply_header(ws, headers, 3)

        sum_amount = 0.0
        sum_tax = 0.0
        sum_total = 0.0
        row_idx = 4
        for i, inv in enumerate(invoices):
            amount = _safe_float(inv.amount)
            tax = _safe_float(inv.tax)
            total = _safe_float(inv.total)
            sum_amount += amount
            sum_tax += tax
            sum_total += total
            source_label = (
                "电子发票"
                if (inv.source_type or "").lower() == "pdf"
                else ("纸质发票" if (inv.source_type or "").lower() == "paper" else (inv.source_type or "-"))
            )
            _apply_data_row(
                ws,
                row_idx,
                [
                    i + 1,
                    inv.invoice_no or "-",
                    inv.invoice_date or "-",
                    inv.seller_name or "-",
                    inv.buyer_name or "-",
                    (inv.items or "").strip() or "-",
                    amount,
                    tax,
                    total,
                    inv.category or "未分类",
                    source_label,
                ],
                money_cols=money_cols,
                is_odd=(i % 2 == 1),
            )
            row_idx += 1

        if not invoices:
            cell = ws.cell(row=row_idx, column=1, value="本月暂无发票数据")
            ws.merge_cells(
                start_row=row_idx,
                start_column=1,
                end_row=row_idx,
                end_column=col_count,
            )
            cell.font = SUBTITLE_FONT
            cell.alignment = CENTER_ALIGN
            row_idx += 1

        # 合计行
        _apply_total_row(
            ws,
            row_idx,
            [
                "合计",
                "",
                "",
                "",
                "",
                f"共 {len(invoices)} 张",
                sum_amount,
                sum_tax,
                sum_total,
                "",
                "",
            ],
            money_cols=money_cols,
        )

        ws.freeze_panes = "A4"
        _autosize_columns(ws, col_count)

        out_path = _build_filename("detail", year, month)
        wb.save(out_path)
        logger.info(f"生成明细Excel: {out_path} (发票数 {len(invoices)})")
        return str(out_path), len(invoices)


# ---------------------------------------------------------------------------
# 文件管理
# ---------------------------------------------------------------------------
def list_export_files() -> List[dict]:
    """列出当前导出目录下所有 xlsx 文件(按时间倒序)"""
    out_dir = ensure_export_dir()
    items: List[dict] = []
    for f in out_dir.iterdir():
        if not f.is_file() or f.suffix.lower() != ".xlsx":
            continue
        try:
            stat = f.stat()
        except OSError:
            continue
        # 解析模式: invoices_{mode}_{year}_{month}_{ts}.xlsx
        mode = None
        parts = f.stem.split("_")
        if len(parts) >= 2 and parts[0] == "invoices":
            mode = parts[1]
        items.append(
            {
                "filename": f.name,
                "size": stat.st_size,
                "created_at": stat.st_mtime,
                "mode": mode,
            }
        )
    items.sort(key=lambda x: x["created_at"], reverse=True)
    return items


def cleanup_old_export_files(days: int = 30) -> int:
    """清理早于N天的导出文件"""
    out_dir = ensure_export_dir()
    cutoff = time.time() - days * 86400
    removed = 0
    for f in out_dir.iterdir():
        if not f.is_file():
            continue
        try:
            if f.stat().st_mtime < cutoff:
                f.unlink()
                removed += 1
        except OSError as e:
            logger.warning(f"清理导出文件失败: {f}, error={e}")
    return removed
